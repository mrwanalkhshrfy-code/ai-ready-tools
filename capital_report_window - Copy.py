import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog
import sqlite3
import os
from datetime import datetime

# مكتبات تصدير Excel والتنسيق
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# مكتبات تصدير PDF والتنسيق
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

import sys # تأكد من إضافة هذه المكتبة في أعلى الملف مع باقي المكتبات

from db_config import get_db_path, get_connection

class CapitalReportWindow(ctk.CTkToplevel):
    def __init__(self, db_path=None):
        super().__init__()
        self.db_path = get_db_path() if not db_path else db_path
        # ... باقي كود التهيئة
        
        # تعريف المتغيرات الأولية
        self.date_str = datetime.now().strftime("%d/%m/%Y")
        self.inventory = 0.0
        self.cash_balance = 0.0
        self.total_assets = 0.0
        self.cust_owed = 0.0
        self.cust_due = 0.0
        self.total_cust = 0.0
        self.supp_owed = 0.0
        self.supp_due = 0.0
        self.total_supp = 0.0
        self.capital = 0.0

        self.title("تقرير رأس المال")
        self.geometry("900x780")
        self.grab_set()

        # ألوان التصميم
        self.COLOR_GRAY = "#EBEBEB"
        self.COLOR_GREEN = "#D5F8D3"
        self.COLOR_ORANGE = "#FFE5D9"
        self.COLOR_WHITE = "#FFFFFF"

        self.setup_ui()
        self.load_data()

    # ... (دالة setup_ui تبقى كما هي بدون تغيير) ...

    
    def setup_ui(self):
        # === شريط التحكم والفلترة بالأعلى ===
        top_bar = ctk.CTkFrame(self, fg_color="transparent")
        top_bar.pack(fill="x", padx=20, pady=15)

        ctk.CTkLabel(top_bar, text="من تاريخ:", font=("Arial", 14, "bold")).pack(side="right", padx=5)
        self.start_date_entry = ctk.CTkEntry(top_bar, width=110, font=("Arial", 13))
        self.start_date_entry.pack(side="right", padx=5)
        self.start_date_entry.insert(0, "2020-01-01")

        ctk.CTkLabel(top_bar, text="إلى تاريخ:", font=("Arial", 14, "bold")).pack(side="right", padx=5)
        self.end_date_entry = ctk.CTkEntry(top_bar, width=110, font=("Arial", 13))
        self.end_date_entry.pack(side="right", padx=5)
        self.end_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

        btn_refresh = ctk.CTkButton(top_bar, text="🔄 تحديث البيانات", font=("Arial", 13, "bold"), fg_color="#2980b9", width=120, command=self.load_data)
        btn_refresh.pack(side="right", padx=15)

        btn_excel = ctk.CTkButton(top_bar, text="📊 تصدير Excel", font=("Arial", 13, "bold"), fg_color="#27ae60", hover_color="#2ecc71", width=120, command=self.export_excel)
        btn_excel.pack(side="left", padx=5)

        btn_pdf = ctk.CTkButton(top_bar, text="📄 تصدير PDF", font=("Arial", 13, "bold"), fg_color="#c0392b", hover_color="#e74c3c", width=120, command=self.export_pdf)
        btn_pdf.pack(side="left", padx=5)

        # === منطقة التقرير الرئيسية ===
        self.report_card = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=10, border_width=1, border_color="#CCCCCC")
        self.report_card.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # 1. عنوان التقرير والتاريخ
        self.title_label = ctk.CTkLabel(self.report_card, text="تقرير رأس المال", font=("Arial", 26, "bold"), text_color="#000000")
        self.title_label.pack(pady=(20, 5))

        self.date_label = ctk.CTkLabel(self.report_card, text=self.date_str, font=("Arial", 22, "bold"), text_color="#000000")
        self.date_label.pack(pady=(0, 15))

        # 2. إطار الجدول الرئيسي (3 أعمدة)
        self.grid_frame = ctk.CTkFrame(self.report_card, fg_color="#CCCCCC")
        self.grid_frame.pack(fill="both", expand=True, padx=25, pady=10)

        for col in range(3):
            self.grid_frame.columnconfigure(col, weight=1, uniform="col")
        for row in range(8):
            self.grid_frame.rowconfigure(row, weight=1, uniform="row")

    def create_cell(self, text, row, col, bg_color, font_size=18, bold=True, columnspan=1):
        """إنشاء خلية بتصميم مطابق"""
        frame = tk.Frame(self.grid_frame, bg=bg_color, highlightbackground="#B0B0B0", highlightthickness=1)
        frame.grid(row=row, column=col, columnspan=columnspan, sticky="nsew", padx=1, pady=1)

        font_weight = "bold" if bold else "normal"
        lbl = tk.Label(frame, text=text, bg=bg_color, fg="#000000", font=("Arial", font_size, font_weight))
        lbl.pack(expand=True, fill="both")
        return lbl

    def load_data(self):
        """حساب أرقام النظام وعرضها"""
        start_d = self.start_date_entry.get() + " 00:00:00"
        end_d = self.end_date_entry.get() + " 23:59:59"
        self.date_str = datetime.now().strftime("%d/%m/%Y")
        self.date_label.configure(text=self.date_str)

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # 1. المخزون
            cursor.execute("SELECT SUM(cost_price * stock_quantity) FROM Products")
            self.inventory = cursor.fetchone()[0] or 0.0

            # 2. الصناديق (النقد الفعلي المستلم من المبيعات وليس الإجمالي)
            # تم التعديل هنا لاستخدام paid_amount بدلاً من total_amount
            cursor.execute("SELECT SUM(paid_amount) FROM Invoices WHERE date_time BETWEEN ? AND ?", (start_d, end_d))
            sales_cash = cursor.fetchone()[0] or 0.0
            
            cursor.execute("SELECT SUM(amount) FROM Expenses WHERE date BETWEEN ? AND ?", (start_d, end_d))
            expenses = cursor.fetchone()[0] or 0.0
            
            self.cash_balance = sales_cash - expenses
            self.total_assets = self.inventory + self.cash_balance

            # 3. العملاء
            cursor.execute("SELECT SUM(balance) FROM Customers WHERE balance > 0")
            self.cust_owed = cursor.fetchone()[0] or 0.0
            cursor.execute("SELECT ABS(SUM(balance)) FROM Customers WHERE balance < 0")
            self.cust_due = cursor.fetchone()[0] or 0.0
            self.total_cust = self.cust_owed - self.cust_due

            # 4. الموردين
            cursor.execute("SELECT ABS(SUM(balance)) FROM Suppliers WHERE balance < 0")
            self.supp_owed = cursor.fetchone()[0] or 0.0
            cursor.execute("SELECT SUM(balance) FROM Suppliers WHERE balance > 0")
            self.supp_due = cursor.fetchone()[0] or 0.0
            self.total_supp = self.supp_owed - self.supp_due

            # 5. رأس المال
            self.capital = self.total_assets + self.total_cust + self.total_supp

            conn.close()
        except Exception as e:
            messagebox.showerror("خطأ", f"خطأ بقاعدة البيانات: {e}")
            return

        # تنظيف وعرض الخلايا
        for widget in self.grid_frame.winfo_children():
            widget.destroy()

        # الصف 1 + 2
        self.create_cell("الإجمالي", 0, 0, self.COLOR_GRAY)
        self.create_cell("رصيد الصناديق", 0, 1, self.COLOR_GRAY)
        self.create_cell("المخزون", 0, 2, self.COLOR_GRAY)

        self.create_cell(f"{self.total_assets:,.0f}", 1, 0, self.COLOR_GREEN, font_size=20)
        self.create_cell(f"{self.cash_balance:,.0f}", 1, 1, self.COLOR_WHITE, font_size=20)
        self.create_cell(f"{self.inventory:,.0f}", 1, 2, self.COLOR_WHITE, font_size=20)

        # الصف 3 + 4
        self.create_cell("الإجمالي", 2, 0, self.COLOR_GRAY)
        self.create_cell("الباقي للعملاء", 2, 1, self.COLOR_GRAY)
        self.create_cell("الباقي عند العملاء", 2, 2, self.COLOR_GRAY)

        self.create_cell(f"{self.total_cust:,.0f}", 3, 0, self.COLOR_GREEN, font_size=20)
        self.create_cell(f"{self.cust_due:,.2f}", 3, 1, self.COLOR_WHITE, font_size=20)
        self.create_cell(f"{self.cust_owed:,.0f}", 3, 2, self.COLOR_WHITE, font_size=20)

        # الصف 5 + 6
        self.create_cell("الإجمالي", 4, 0, self.COLOR_ORANGE)
        self.create_cell("الباقي للموردين", 4, 1, self.COLOR_ORANGE)
        self.create_cell("الباقي عند الموردين", 4, 2, self.COLOR_ORANGE)

        self.create_cell(f"{self.total_supp:,.0f}", 5, 0, self.COLOR_WHITE, font_size=20)
        self.create_cell(f"{self.supp_due:,.0f}", 5, 1, self.COLOR_WHITE, font_size=20)
        self.create_cell(f"{self.supp_owed:,.2f}", 5, 2, self.COLOR_WHITE, font_size=20)

        # الصف 7 + 8
        self.create_cell("رأس المال", 6, 0, self.COLOR_GREEN, font_size=24, columnspan=3)
        self.create_cell(f"{self.capital:,.0f}", 7, 0, self.COLOR_GREEN, font_size=28, columnspan=3)
    # ==================== تصدير EXCEL ====================
    def export_excel(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="حفظ تقرير رأس المال")
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير رأس المال"
        ws.views.sheetView[0].rightToLeft = True

        fill_gray = PatternFill("solid", fgColor="EBEBEB")
        fill_green = PatternFill("solid", fgColor="D5F8D3")
        fill_orange = PatternFill("solid", fgColor="FFE5D9")

        font_title = Font(name="Arial", size=18, bold=True)
        font_header = Font(name="Arial", size=14, bold=True)
        font_val = Font(name="Arial", size=14, bold=True)
        font_capital = Font(name="Arial", size=20, bold=True)

        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
                             top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'))

        ws.merge_cells("A1:C1")
        ws["A1"] = "تقرير رأس المال"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_center

        ws.merge_cells("A2:C2")
        ws["A2"] = self.date_str
        ws["A2"].font = font_header
        ws["A2"].alignment = align_center

        grid_data = [
            [("الإجمالي", fill_gray), ("رصيد الصناديق", fill_gray), ("المخزون", fill_gray)],
            [(f"{self.total_assets:,.0f}", fill_green), (f"{self.cash_balance:,.0f}", None), (f"{self.inventory:,.0f}", None)],

            [("الإجمالي", fill_gray), ("الباقي للعملاء", fill_gray), ("الباقي عند العملاء", fill_gray)],
            [(f"{self.total_cust:,.0f}", fill_green), (f"{self.cust_due:,.2f}", None), (f"{self.cust_owed:,.0f}", None)],

            [("الإجمالي", fill_orange), ("الباقي للموردين", fill_orange), ("الباقي عند الموردين", fill_orange)],
            [(f"{self.total_supp:,.0f}", None), (f"{self.supp_due:,.0f}", None), (f"{self.supp_owed:,.2f}", None)],
        ]

        start_row = 4
        for r_idx, row in enumerate(grid_data):
            for c_idx, (val, fill) in enumerate(row):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=val)
                cell.font = font_header if r_idx % 2 == 0 else font_val
                cell.alignment = align_center
                cell.border = thin_border
                if fill: cell.fill = fill

        # دمج خلايا رأس المال
        ws.merge_cells("A10:C10")
        ws["A10"] = "رأس المال"
        ws["A10"].font = font_capital
        ws["A10"].fill = fill_green
        ws["A10"].alignment = align_center
        ws["A10"].border = thin_border

        ws.merge_cells("A11:C11")
        ws["A11"] = f"{self.capital:,.0f}"
        ws["A11"].font = font_capital
        ws["A11"].fill = fill_green
        ws["A11"].alignment = align_center
        ws["A11"].border = thin_border

        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 25

        wb.save(filepath)
        messagebox.showinfo("نجاح", "تم تصدير ملف Excel بنجاح بنفس التنسيق!")

    # ==================== تصدير PDF ====================
    def get_registered_arabic_font(self):
        """البحث عن خط عربي ومسجل بالنظام لمنع ظهور رموز التشفير"""
        font_paths = [
            "C:\\Windows\\Fonts\\arial.ttf",
            "C:\\Windows\\Fonts\\tahoma.ttf",
            "C:\\Windows\\Fonts\\seguiemj.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        ]
        for path in font_paths:
            if os.path.exists(path):
                try:
                    font_name = "ArabicSystemFont"
                    pdfmetrics.registerFont(TTFont(font_name, path))
                    return font_name
                except Exception:
                    continue
        return None

    def export_pdf(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")], title="حفظ تقرير PDF")
        if not filepath:
            return

        # 1. التحقق مما إذا كان الملف مفتوحاً في مكان آخر قبل التعديل عليه
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except PermissionError:
                messagebox.showerror("خطأ في الوصول", "الملف مفتوح حالياً في المتصفح أو برنامج آخر.\nيرجى إغلاقه أولاً ثم إعاده التصدير.")
                return

        font_name = self.get_registered_arabic_font()
        if not font_name:
            messagebox.showerror("خطأ", "لم يتم العثور على خط عربي مثبت بالنظام (مثل Arial أو Tahoma).")
            return

        def ar(text):
            if text is None: return ""
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)

        try:
            doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            elements = []

            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'], fontName=font_name, fontSize=22, alignment=1, leading=26)
            date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontName=font_name, fontSize=16, alignment=1, leading=20)
            
            cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName=font_name, fontSize=12, alignment=1, leading=14)
            cell_bold = ParagraphStyle('CellBold', parent=styles['Normal'], fontName=font_name, fontSize=13, alignment=1, leading=15)
            capital_bold = ParagraphStyle('CapBold', parent=styles['Normal'], fontName=font_name, fontSize=18, alignment=1, leading=22)

            elements.append(Paragraph(ar("تقرير رأس المال"), title_style))
            elements.append(Spacer(1, 8))
            elements.append(Paragraph(ar(self.date_str), date_style))
            elements.append(Spacer(1, 20))

            table_data = [
                [Paragraph(ar("الإجمالي"), cell_bold), Paragraph(ar("رصيد الصناديق"), cell_bold), Paragraph(ar("المخزون"), cell_bold)],
                [Paragraph(f"{self.total_assets:,.0f}", cell_bold), Paragraph(f"{self.cash_balance:,.0f}", cell_style), Paragraph(f"{self.inventory:,.0f}", cell_style)],
                
                [Paragraph(ar("الإجمالي"), cell_bold), Paragraph(ar("الباقي للعملاء"), cell_bold), Paragraph(ar("الباقي عند العملاء"), cell_bold)],
                [Paragraph(f"{self.total_cust:,.0f}", cell_bold), Paragraph(f"{self.cust_due:,.2f}", cell_style), Paragraph(f"{self.cust_owed:,.0f}", cell_style)],
                
                [Paragraph(ar("الإجمالي"), cell_bold), Paragraph(ar("الباقي للموردين"), cell_bold), Paragraph(ar("الباقي عند الموردين"), cell_bold)],
                [Paragraph(f"{self.total_supp:,.0f}", cell_style), Paragraph(f"{self.supp_due:,.0f}", cell_style), Paragraph(f"{self.supp_owed:,.2f}", cell_style)],
                
                [Paragraph(ar("رأس المال"), capital_bold), "", ""],
                [Paragraph(f"{self.capital:,.0f}", capital_bold), "", ""]
            ]

            col_w = 170
            t = Table(table_data, colWidths=[col_w, col_w, col_w], rowHeights=[32]*8)
            
            t.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#B0B0B0")),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#B0B0B0")),

                ('BACKGROUND', (0, 0), (2, 0), colors.HexColor("#EBEBEB")),
                ('BACKGROUND', (0, 1), (0, 1), colors.HexColor("#D5F8D3")),

                ('BACKGROUND', (0, 2), (2, 2), colors.HexColor("#EBEBEB")),
                ('BACKGROUND', (0, 3), (0, 3), colors.HexColor("#D5F8D3")),

                ('BACKGROUND', (0, 4), (2, 4), colors.HexColor("#FFE5D9")),

                ('SPAN', (0, 6), (2, 6)),
                ('BACKGROUND', (0, 6), (2, 6), colors.HexColor("#D5F8D3")),

                ('SPAN', (0, 7), (2, 7)),
                ('BACKGROUND', (0, 7), (2, 7), colors.HexColor("#D5F8D3")),
            ]))

            elements.append(t)
            doc.build(elements)

            messagebox.showinfo("نجاح", "تم تصدير ملف PDF بنجاح!")
            
            # فتح الملف تلقائياً فور حفظه
            import platform, subprocess
            if platform.system() == "Windows":
                os.startfile(filepath)
            elif platform.system() == "Darwin":
                subprocess.call(["open", filepath])
            else:
                subprocess.call(["xdg-open", filepath])

        except Exception as e:
            messagebox.showerror("خطأ أثناء التصدير", f"تعذر إنشاء ملف PDF بسبب:\n{str(e)}")